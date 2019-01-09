""" Python cheat sheet
        - A collection of python code snippets
        

"""

#============================================================================
#%% Libraries
#============================================================================

"""
# Libraries
# ==================

# Update a library after changing it
import myLib
reload(myLib) # Probably best to do from command line

In Python 3
import imp
imp.reload(module)

e.g if imported mylib using
import mylib as mb

you would reload using
imp.reload(mb)
"""

#============================================================================
#%% Files
#============================================================================

# Reading a csv file using numpy
from numpy import genfromtxt
data = genfromtxt(csvfile, delimiter=',', dtype=None)

# Returned data is a recarray

# Read CIBC .csv file and skip the lines where there are only 2
# columns
# Note: This is because genfromtxt is crap at handling missing values
data = genfromtxt(csvfile, delimiter=',', dtype=None,filling_values=0,names = 'A,B,C,D',invalid_raise=False)

# Finding files on the python search path
import os
print(os.path.abspath("numpy.py"))


# Current working directory
import os
os.getcwd()

# Get home directory of current user
os.path.expanduser('~')

# Change working directory
import os
os.chdir("/home/redlegjed/John/Drawing/GimpScripts")

# Get a directory listing
os.listdir()

# Check if directory exists
os.path.isdir(fpath)

# Check if file/path exists
os.path.exists(fpath)

# Extract filename from a full path/filename
f = "/home/redlegjed/some.txt"
os.path.basename(f)

# Extract path and filename from a full path/filename
path, filename = os.path.split(f)

# Split extension from name
name,ext = os.path.splitext("some.txt")

# Joining a path and a filename
os.path.join("/home","John")

# Making directories
os.mkdir(dirname) # Make single directory

os.makedirs(path)  # Make all paths required to make the complete path

# delete a file
os.remove(path_and_filename)


#%% Finding files
# Help gives the path at the end
help("<name of file>")
 
 # Adding to the system path
import sys
sys.path.append('abspath')

# or
sys.path.append('relpath')

#or
sys.path.insert(0,modulePath)

# Changing the coding of a text file programmatically
# pasted from Stack Overflow, not tested yet

# May need to have imported this
import codecs

# Code snippet
sourceEncoding = "iso-8859-1"
targetEncoding = "utf-8"
source = open("source")
target = open("target", "w")

target.write(unicode(source.read(), sourceEncoding).encode(targetEncoding))

#============================================================================
#%% Running system commands
#============================================================================
import subprocess

cmd = 'ls -l'

return_arg = subproces.call(cmd)

# Not verified yet

#============================================================================
#%% CSV files
#============================================================================


import csv
with open('some.csv', newline='') as f:
    reader = csv.reader(f)
    for row in reader:
        print(row)
  
#============================================================================
#%% Debugging
#============================================================================      


# Using ipython import the embed module
from IPython import embed

# Place an embed() command where you want to pause the program
# this will open an ipython console at that point and allow you to
# see variables and edit them
embed()

#============================================================================
#%% Variable/Class information
#============================================================================

# What class is a variable
type(variable)

# Methods in a module/class
# only if the class has a __dir__().
dir('ClassName')

# inspect module has a lot of is* type functions
import inspect

# If a class is defined as:
class MyClass:
    """A simple example class"""
    
    # Class variables defined here
    # These are COMMON to ALL instances of the class created
    var1 = "Every instance will see me"
    var2 = 3
    
    def __init__(self):
        # Define any standard variables that are SPECIFIC to the instance here
        self.i = 12345
        self.f()
        
    def f(self):
        print(self.i)
        return 'hello world'
        
# An instance of the class is created by
x = MyClass()
# the brackets are important otherwise it doesn't working
#if its at the end of a program,do this:
if __name__ == "__main__":
    x = MyClass()
    
# if you want to make your class a sub class,do this:

class MyClass(className):
    ...

#============================================================================
#%% Numpy
#============================================================================
 
 # Numpy arrays
 # 3x2 array of zeros
np.zeros((3,2))
 
 # initialise array and specify the column types
 # don't specify number of columns - leave that to dtype
a = np.zeros((3,),dtype=[('x',float),('y',int),('z',str)])
 
 # Create an array directly from numbers
 x = np.array([[1,2,5],[3,4,6]])
 #
 # x = 1 2
 #     3 4

# Create a column vector
# literally put newaxis in, it's not a variable
x = np.arange(0,10)[:,np.newaxis]

# Convert scalar to array
x = 4.3
x = np.atleast_1d(x)  # Convert to array with at least one dimension
 
 # Random integers
 # 0 - 9
 np.random.randint(10)
 
 # Array indexing
 # 2nd row 3rd column
 a[2][3]
 
 # All 2nd row
 a[2]
 
 # All 1st column
 a[:,0]
 
 # All 2nd column
 a[:,1]
 
 # Combining two vectors into a two column array
 a = np.arange(1,11)
 np.column_stack((a,a))
 
 # Stacking 2 column arrays
 a = np.zeros((1,2))
 b = np.ones((1,2))
 np.concatenate((a,b))
 
 # Converting float array to integers
 a.astype(int)
 
 # Mix of string and numeric colums
 # This example is reading a file that has format [Num String Num Num]
 # Create array first    
 # Note the 'a20' for strings 20 characters long
    data = np.zeros((len(txtLines)),
                      dtype = [('Date',datetime.datetime),
                               ('Transaction','a20'),
                                ('Debit',float),
                                ('Credit',float)])

# Find ranges for numpy data types
# ----------------------------------

# floats
np.finfo(np.float32)

# integers
np.iinfo(np.int16)

# Convert structured array to normal array
# x = structured array
x.view((float, len(x.dtype.names)))

# Boolean operations
# --------------------
# e.g. removing NaNs
x = x[~numpy.isnan(x)]
# or
x = x[numpy.isfinite(x)]

# Get Max-Min (i.e. Peak-to-Peak) of array
x = np.random.rand(10)
x.ptp() # Min - Max

#============================================================================
#%% Strings
#============================================================================ 

# Convert bytes object to string
string = bytesObject.decode("utf-8")
# or
string = str(bytesObject,"utf-8")

# string to bytes
bytesObject = string.encode()

# Split into list
strVar.split(",")

# strip characters from front and back of string
# in this case strip off * characters from each end
strVar.strip("*")

# Convert to numeric
float(strVar)
int(strVar)

# Get alphabet
import string

lowerCaseLetters = string.ascii_lowercase 
upperCaseLetters = string.ascii_uppercase 
lowerPlusUpper = string.ascii_letters

#============================================================================
#%% Dictionaries
#============================================================================

# How to make one:
d = {"string": "ROGERS","limit": 70,"indices":[]}
# How to read one
d['string']

#how to add to one
d["YAHHO"]="not here"

# Remove item if only key is known
d.pop(keyString,None)

# How to loop through keys and values
for key,value in my_dict.items():
    print(key,'=',value)


#============================================================================
#%% Tuples
#============================================================================

from collections import namedtuple

Point = namedtuple('Point','x_coord y_coord')
my_point = Point(12,4)
print('My Point is at (%i,%i)' % (my_point.x_coord,my_point.y_coord))

#============================================================================
#%% Lists
#============================================================================

# Convert list to string
# 1. With no delimiters
ll = ['a','b','c']
string = ''.join(ll)

# 2. with delimiters e.g. carriage returns
delim = '\n'.join(ll)
print(delim)

# Applying a function to all elements in a list
# This is called a "list comprehension" in Python-speak
# 
# the example applies the len() function to every item in the list
num = [len(x) for x in ETF_list]

# Finding unique elements in a list
# ----------------------------------
def unique(l):
    s = set(); n = 0
    for x in l:
        if x not in s: s.add(x); l[n] = x; n += 1
    del l[n:]

#============================================================================
#%% Matplotlib
#============================================================================

# Get path where Matplotlib stores its data
import matplotlib as mpl
mpl.get_data_path()


import matplotlib.pyplot as plt
#matplotlib.use('Cairo')
# plt.ion()
plt.plot([1,2,3,4])
plt.ylabel('some numbers')
plt.show()

# Set size of figure
# ---------------------
# size is in inches
figure(num=None, figsize=(8, 6), dpi=80, facecolor='w', edgecolor='k')

# Set name of figure window
fig = plt.gcf()
fig.canvas.set_window_title('Test')

# Setting legend font size to 8
legend(prop={'size':8})

# subplot spacing
# -------------------
f = figure(1)

# Set subplot spacing
f.subplots_adjust(left=None, bottom=None, right=None, top=None,
                    wspace=None, hspace=None)

# Get spacing on figure
f.subplotpars.bottom
f.subplotpars.top
# etc, i.e. all parameters from subplots_adjust

# Minor ticks
# ------------------
minorLocator   = AutoMinorLocator()
ax = plt.gca()
ax.xaxis.set_minor_locator(minorLocator)
ax.yaxis.set_minor_locator(minorLocator)

#or 
plt.minorticks_on()

# Plotting dates
# ------------------

# import this
import matplotlib.dates as mdates

# Do this to the plot
ax = plt.gca()
fig.autofmt_xdate()
ax.fmt_xdata = mdates.DateFormatter('%Y-%m-%d')
fig.show()

# Adding arrows
# Note: first parameter is empty string
ax.annotate("",
            xy=(0.2, 0.2), xycoords='data',
            xytext=(0.8, 0.8), textcoords='data',
            arrowprops=dict(arrowstyle="->",
                            connectionstyle="arc3"),
            )
            
# Using latex
# Put the latex commands in between $ signs
title("Some math text $\alpha_i + \beta^2$")

# Display current figure inline
display(plt.gcf())

# Customising defaults
# http://matplotlib.org/users/customizing.html

# rc file location
import matplotlib
matplotlib.matplotlib_fname()

#%% Set figure style

import matplotlib.pyplot as plt

# Set the style
plt.style.use('ggplot')

# Get available styles
print(plt.style.available)

# ['dark_background', 'ggplot', 'bmh', 'fivethirtyeight', 'grayscale']


#%% Make figures with white background
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
import matplotlib as mpl

mpl.rcParams["figure.facecolor"] = 'white'
mpl.rcParams["figure.edgecolor"] = 'black'
mpl.rcParams["axes.facecolor"] = 'white'
mpl.rcParams["axes.edgecolor"] = 'black'
mpl.rcParams["grid.color"] = 'black'

#%% Make figures with black background
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
import matplotlib as mpl

mpl.rcParams["figure.facecolor"] = 'black'
mpl.rcParams["figure.edgecolor"] = 'white'
mpl.rcParams["axes.facecolor"] = 'black'
mpl.rcParams["axes.edgecolor"] = 'white'
mpl.rcParams["grid.color"] = 'white'

#%% Setting the colour cycle for plots

fig = plt.figure(figname)
plt.clf()
ax = plt.gca()
ax.set_color_cycle([plt.cm.Accent(i) for i in np.linspace(0, 1, NumberOfLines)])


#%% Setting the background colour for saved plots

# facecolor = background
# edgecolor = border 
# (note by default the border is zero, it has to be set when creating the 
# figure)

# Set border
fig = plt.figure(linewidth=2)
plt.savefig(filename,facecolor='#2E2E2E',edgecolor='#848484',transparent=True)

#============================================================================
#%% Time/Date
#============================================================================

import time

# Current date/time in ascii format
time.asctime()
#wait
time.sleep(seconds)

# anonymous function to convert date
convDate = lambda d: datetime.strptime(d, '  %Y %m %d %H %M %S')

# Convert datetime object to string with arbitrary format:
newDateString = date_object.strftime("%Y %m %d")

# Convert to ISO format with space between date and time
isoDateString = date_object.isoformat(" ")

# Converting from string to datetime
from datetime import datetime
date_object = datetime.strptime('Jun 1 2005  1:33PM', '%b %d %Y %I:%M%p')
date_object = datetime.strptime('  2012 11 09 10 49 12', '  %Y %m %d %H %M %S')

# Time periods
from datetime import timedelta

dt = timedelta(weeks=1,minutes=3)

# Converting output of os.stat() to readable format
# the index '-2' is picking out gives the time modified
time.strftime('%d/%m/%Y %H:%M:%S', time.localtime(os.stat(some_filename).st_mtime))


#============================================================================
#%% timeit library
#============================================================================

import timeit

# Time a stand alone snippet of code
timeit.timeit("max(range(20))",number=100)

# Time a snippet that uses the global namespace
timeit.timeit("[n for n in a_global_variable]",number=100,globals=globals())


#============================================================================
#%% Exceptions
#============================================================================

# Example of how to get error message from exception
def catch():
     try:
         raise BaseException()
     except Exception as e:
         print( e.message, e.args)

#============================================================================
#%% zip files
#============================================================================

import zipfile

testfile = "/home/redlegjed/Downloads/pgu-0.18.zip"

# Open zip file for reading
zipObj = zipfile.ZipFile(testfile,"r")


# Print list of files
zipObj.printdir()


# Get list of files as zipInfo objects
# - need this for extracting
fnames = zipObj.infolist()


# Extract one file
# to new path
zipObj.extract(fnames[0],path = newPathString)

# close
zipObj.close()


#============================================================================
#%% Web access
#============================================================================


storeDir = "/home/redlegjed/John/Python/datafiles/CH1.PDF"
#URL = "http://www.python.org"
URL = "http://www.dspguide.com/CH1.PDF"

urllib.urlretrieve(URL,storeDir)


"""
Downloading web pages using pyCurl
===============================
Tested and it works
"""

import pycurl
c = pycurl.Curl()
c.setopt(pycurl.URL, "http://www.python.org/")
c.setopt(pycurl.HTTPHEADER, ["Accept:"])
import StringIO
b = StringIO.StringIO()
c.setopt(pycurl.WRITEFUNCTION, b.write)
c.setopt(pycurl.FOLLOWLOCATION, 1)
c.setopt(pycurl.MAXREDIRS, 5)
c.perform()
print b.getvalue()
import urllib

""" 
Downloading files using pyCurl
===============================
untested - took it as an example of how to direct curl output
to a file
"""

import re
import urllib2
import pycurl

url = "http://server.domain/"
path = "path/"
pattern = '<A HREF="/%s.*?">(.*?)</A>' % path

response = urllib2.urlopen(url+path).read()

for filename in re.findall(pattern, response):
    fp = open(filename, "wb")
    curl = pycurl.Curl()
    curl.setopt(pycurl.URL, url+path+filename)
    curl.setopt(pycurl.WRITEDATA, fp)
    curl.perform()
    curl.close()
    fp.close()
    




#============================================================================
#%% Reading JPEG files
#============================================================================


import os,sys
from PIL import Image
from PIL.ExifTags import TAGS

for (k,v) in Image.open(filename)._getexif().iteritems():
        print '%s = %s' % (TAGS.get(k), v)


#============================================================================
#%% Images <-> numpy/matplotlib
#============================================================================


# Convert numpy array im to PIL image
# ------------------------------------
# normalise array to values 0-1 first
im_norm = im/255
# convert to PIL image, using binary_r colormap
# other colormaps can be used
im_pil = PIL.Image.fromarray(np.uint8(cm.binary_r(im_norm)*255))


#============================================================================
#%% Reading .mat files
#============================================================================


# From Stack overflow
# using h5py library : www.h5py.org

import h5py
f = h5py.File('test.mat')
f.keys()

# access data using keys like a dictionary

import numpy as np, h5py 
f = h5py.File('somefile.mat','r') 
data = f.get('data/variable1') 
data = np.array(data) # For converting to numpy array


#============================================================================
#%% Making a Python console
#============================================================================

import code,sys
#make the main console
try:
    raise None
except:
        frame = sys.exc_info()[2].tb_frame.f_back
#bring in stuff you need to make a console
namespace = frame.f_globals.copy()
namespace.update(frame.f_locals)
#create the mane console
code.interact(local=namespace)

###the core of a console...###
#importes
import traceback
import sys
#setup the locals
_locals = {}
def check(txt):
    #make the out put
    _stdout = sys.stdout
    s = sys.stdout = writer() # writer is a class with a 'write(data)'
    val = txt
    if val == "":
        #just do nothing to stop Error
        jgjdfbgshdfg="sbcfhvjasvfbS"
    else:
        
       
        try:
            # compile the code
            code = compile(val,'<stdout>','single')
            #run the code
            eval(code,globals(),_locals)
        except:
            #if there is an error,here is what happens
            
            # get the type,the value and the traceback
            e_type,e_value,e_traceback = sys.exc_info()
            
            #print will print on your commandline
            print('Traceback (most recent call last):')
            
            #this line makes the prints print on your commandline,and writes down the traceback 
            traceback.print_tb(e_traceback,None,s)
            
            # show the type
            print(e_type)
            
            #show the value
            print(e_value)
            

    

    sys.stdout = _stdout
    root.title("python shell")
    #error  = False
    

#============================================================================
#%% Reading Matlab files (version <7.3)
#============================================================================    

 
import os
import scipy.io as sio

os.chdir(r"C:\Users\jbainbri\Documents\0 Projects\BlobData\WL3 MB func results\MatFiles_with_station\NNTMRT0003DW")

mfilename = r"NTTST114_2013-01-02-11h-19m-59s_RxBandwidthVerification_130102_120145_MATFILE_results.mat"

# Read file
mat = sio.loadmat(mfilename)
# this returns a dict where each key is the name of the variables
# stored in the .mat file

# Extract variable from the dict
results = mat['results']

# if results is a Matlab struct then a structured NDarray is returned
# where the fields are dtypes of the NDarray.

# access the fields by:
SNR = results['SNR_H_Rx_dB']
# This gives a ND array

# access contents of SNR using [0,0]
# using index [0] doesn't work for some reason
SNR[0,0]

# if SNR is an array then access individual elements with extra indices
SN[0,0][0,0] # first element
SN[0,0][0,1] # second element



# See which variables are stored in file, same as matlab who command
sio.whosmat(mfilename)


    

#======================================================================
#%% Making a PDF file with Matplotlib
#======================================================================


#%% This is a demo of creating a pdf file with several pages.
# it doesn't work, but gives the idea. The pdf writer can work.

import datetime
import numpy as np
from matplotlib.backends.backend_pdf import PdfPages
import matplotlib.pyplot as plt

# Create the PdfPages object to which we will save the pages:
# The with statement makes sure that the PdfPages object is closed properly at
# the end of the block, even if an Exception occurs.

    
pdf = PdfPages('/home/john/Documents/Wokr/CoRx/data/multipage_pdf.pdf')
fig1 = plt.figure(1)
plt.plot(range(7), [3, 1, 4, 1, 5, 9, 2], 'r-o')
#plt.title('Page One')
pdf.savefig(fig1)  # saves the current figure into a pdf page
plt.close()



# We can also set the file's metadata via the PdfPages object:
d = pdf.infodict()
d['Title'] = 'Multipage PDF Example'
d['Author'] = u'Jouni K. Sepp\xe4nen'
d['Subject'] = 'How to create a multipage pdf file and set its metadata'
d['Keywords'] = 'PdfPages multipage keywords author title subject'
d['CreationDate'] = datetime.datetime(2009, 11, 13)
d['ModDate'] = datetime.datetime.today()

pdf.close()

#============================================================================
#%% Basic PANDAS
#============================================================================


# Indexing dataframe using numeric indices
# rows 0-5, columns 0-2
df.iloc[0:4,0:2]

# Slice out a column
col = df[colNameString]

# Slice using column name
# single column
df1.loc[0:3,col1NameString]

# Multiple columns
df1.loc[0:3,(col1NameString,col2NameString)]

# Get last row
df1.tail(1)
df1['col name'].tail(1)

# Get last 10 rows
df1.tail(10)
df1['col name'].tail(1)

# Get first row(s)
df1['col name'].head(1)



# Get column names
# - various versions of this, all are different types

# Return as pandas.core.index.Index type
df.columns

# Return as numpy.ndarray
df.columns.values
# or
df[key].values
# or to combine several keys and return as 2D array
df[keyList].values

# Return as a list
# - note this one is a function
df.columns.values.tolist()

# Using .loc
df.loc[index_range,column_list]

# e.g.
data.loc[index_variable,['SerialNumber','IsRetest']]

# Check if column is a number
np.issubdtype(df['A'].dtype, np.number)

#%% Load data file into PANDAS dataframe [Windows]
#----------------------------------------------------

# Read in data with dates in some columns
df = pd.read_csv(filename, parse_dates = ["dateColumn1","dateColumn2"])


# Indexing a single or slice of a dataframe from one column
dt.loc[0:2,"Serial Number: "]

# Indexing multiple rows and more than one column
dt.loc[ 0:2, ["Serial Number: ","Test Started at: "] ]


#%% Convert Python 2 saved .csv to Python 3 .csv using Pandas
from io import StringIO

python2_csv = r'C:\Users\jbainbri\Documents\0 Projects\BlobData\WL3 MB func results\TestDataSpreadsheets\WL3_MBfunct 2014-01-01 to 2014-11-30_C.csv'

# Read Python 2 file in as bytes
with open(python2_csv,'rb') as f:
    txt_b = f.read()
   
# decode but ignore errors
txt = txt_b.decode(errors='ignore')

# Convert to file like object
csv = StringIO(txt)

# Read into Pandas
df = pd.read_csv(csv,parse_dates = datetime_columns)

# save from Pandas
df.to_csv(data_filename3)

#%% Pandas date ranges
# -------------------------------

# Between dates
# assuming there is a column called 'date' with pandas timestamp objects in
filtered = df[df['dates'] > '2014-01-23']
filtered = df[(df['dates'] > '2014-01-23') & df['dates'] < '2014-06-06')]

# Make 'dates' column the index
# note only the newly created dataframe has the new index
new_df = df.set_index('dates')

# Now the dataframe can be slices using dates
slice_data = df["2014-01-12":"2014-05-09"]

# Convert time delta to integers
# ---------------------------------
# Calculate time delta by subtraction
dataframe['TotalTestTime_min'] = dataframe.StopTime - dataframe.StartTime

# Convert to minutes
# using np.timedelta64(1,'m') as the reference unit, it's the 'm' that converts
# to minutes, could use 'D' for days, 'h'=hours, 's'=seconds
dataframe['TotalTestTime_min'] = (dataframe.TotalTestTime_min / np.timedelta64(1, 'm')).astype(int)

#%% Pandas filtering columns using regex

data = # a dataframe

P1Channels = data.filter(regex="P1")
P1Sum = P1Channels.sum(axis=1)

#%% Pandas plotting
# ---------------------

# Plot one column against another as a line plot
df.plot(x='A', y='B')

# Scatter plot
df.plot(x='A', y='B',kind='scatter')

#==========================================================================
#%% How to get the arguments from a function [stackoverflow]
#==========================================================================

import inspect

def thefunction(pos1, pos2, a=1,b=2,c=3, *args, **kwargs):
    pass

argspec = inspect.getargspec(thefunction)

print(argspec)
# ArgSpec(args=['pos1', 'pos2', 'a', 'b', 'c'], varargs='args', keywords='kwargs', defaults=(1, 2, 3))

print(argspec.args)
# ['pos1', 'pos2', 'a', 'b', 'c']

print(argspec.args[-len(argspec.defaults):])
# ['a', 'b', 'c']


#==========================================================================
#%% Dynamically adding functions to a class
#==========================================================================

http://www.ianlewis.org/en/dynamically-adding-method-classes-or-class-instanc

#==========================================================================
#%% Download big files from web
#==========================================================================
http://stackoverflow.com/questions/7243750/download-file-from-web-in-python-3


#===========================================================================
#%% Unique colours
#===========================================================================

#%% From stack overflow
import numpy as np
import colorsys

def _get_colors(num_colors):
    colors=[]
    for i in np.arange(0., 360., 360. / num_colors):
        hue = i/360.
        lightness = (50 + np.random.rand() * 10)/100.
        saturation = (90 + np.random.rand() * 10)/100.
        colors.append(colorsys.hls_to_rgb(hue, lightness, saturation))
    return colors
#%%
    
#===========================================================================
#%% Sympy
#===========================================================================

# Setup for black background
from sympy import init_printing
from sympy import init_session

init_printing(forecolor="White") # Set font colour
init_session()

#===========================================================================
#%% iPython
#===========================================================================
# How to detect it
def in_ipython():
    try:
        __IPYTHON__
    except NameError:
        return False
    else:
        return True

#===========================================================================
#%% Hex<=>Dec
#===========================================================================

def dec2hex(n):
    """return the hexadecimal string representation of integer n"""
    return "%X" % n
    
def hex2dec(s):
    """return the integer value of a hexadecimal string s"""
    return int(s, 16)
    
#============================================================================
#%% Read from Excel spreadsheet
#============================================================================
from openpyxl import load_workbook

wb = load_workbook(spreadsheet,data_only=True)
ws = wb.get_sheet_by_name('sheetname')

ss_list = []

for row in ws['A8':'A43']:
    for cell in row:
        print(cell.value)
        ss_list.append(cell.value)


#============================================================================
#%% Read XML file
#============================================================================
from bs4 import BeautifulSoup

with open(xml_file,'r') as f:
    xml = f.read()
    
soup = BeautifulSoup(xml,'xml')

# Get readings
reading = soup.findAll('Reading')

xml_list = []

# Print out list of readings
for result in reading:
    print("%s : %s" % (result['Name'],result['Value']))
    xml_list.append(result['Name'])
